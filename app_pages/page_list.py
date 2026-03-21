# ============================================
# Page List
# ============================================

# Importar librerías 
import streamlit as st
import pandas as pd
import os
from PIL import Image
from openpyxl import load_workbook
import base64
from fpdf import FPDF
import streamlit.components.v1 as components
from io import BytesIO

# Importar rutas de configuración
from common.config import (
    ASSETSIMG, DATA_DIR, ASSETSFONTS
)

# Importar funciones de datos, identificación de usuarios y marca de agua
from controllers.db_controller import load_stats_players_fbref
from controllers.user_controller import load_users
from common.pdf_utils import get_watermark
    
# Función que da cómo resultado la página List
def page_list():
    
    # Añadir CSS de impresión
    st.markdown(
    """
    <style>

    /* estilos solo cuando se imprime */
    @media print {
        
        body {
            zoom: 0.7;
        }

        /* ocultar sidebar */
        section[data-testid="stSidebar"] {
            display: none;
        }

        /* ocultar menú superior de streamlit */
        header {
            display: none;
        }

        /* ocultar footer */
        footer {
            display: none;
        }

        /* expandir contenido principal */
        .main {
            margin-left: 0px;
            width: 100%;
        }

    }

    </style>
    """,
    unsafe_allow_html=True
    )
    
    # Se crean 3 columnas para centrar la imagen
    list_imgcol1, list_imgcol2, list_imgcol3 = st.columns([3, 2, 3])

    with list_imgcol2:

        # Se carga una imagen desde la carpeta assets
        list_img_header = Image.open(ASSETSIMG / "Freepick List.png")
        
        # Se muestra la imagen en Streamlit
        st.image(list_img_header, width=120)

    # Título de la página
    st.header("Player List")
    
    # Cargar datos desde la DB
    list_df_players = load_stats_players_fbref()
    
    # Cargar usuarios como objetos
    list_users = load_users()
    list_users_list = [u.username for u in list_users]  
    
    # Constantes
    LISTS = ["GK", "DF", "LT", "MF", "EX", "FW"]  # Tipos de lista/posiciones
    EXCEL_FILE = DATA_DIR / "list_players_register.xlsx"
    pos_order = ["All","GK","DF","MF","FW"]
    
    # Session state defaults
    list_defaults = {

        "list_league_filter": "All",
        "list_team_filter": "All",
        "list_pos_filter": "All",
        "list_player_to_add": "Select",

        "list_user_select": "Select",
        "list_list_choice": "All",
        "list_comment": "",
        "list_note": "Target",

        "list_view_list": "All",
        "list_view_users_list": "All",

        "list_do_reset": False
    }
    
    for key, value in list_defaults.items():
        st.session_state.setdefault(key, value)
    
    # Reset inteligente
    if st.session_state["list_do_reset"]:

        st.session_state["list_league_filter"] = "All"
        st.session_state["list_team_filter"] = "All"
        st.session_state["list_pos_filter"] = "All"
        st.session_state["list_player_to_add"] = "Select"

        st.session_state["list_user_select"] = "Select"
        st.session_state["list_list_choice"] = "All"
        st.session_state["list_comment"] = ""
        st.session_state["list_note"] = "Target"

        st.session_state["list_view_list"] = "All"
        st.session_state["list_view_users_list"] = "All"

        st.session_state["list_do_reset"] = False

        st.rerun()

    # Función auxiliar para cargar hoja de Excel
    def load_sheet(sheet_name):

        # Columnas esperadas en el archivo
        columns = ["Player", "User", "League", "Team", "Position", "List", "Comment", "Note"]

        # Si el archivo no existe, devolver DataFrame vacío
        if not os.path.exists(EXCEL_FILE):
            return pd.DataFrame(columns=columns)

        try:
            # Intentar cargar la hoja solicitada
            return pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        except:
            # Si la hoja no existe o falla la lectura, devolver DataFrame vacío
            return pd.DataFrame(columns=columns)

    # Función auxiliar para guardar jugador en Excel
    def save_player_to_excel(data, sheet_name):

        # Convertir los datos del jugador a DataFrame
        df_new = pd.DataFrame([data])

        # Si el archivo no existe, crear uno nuevo
        if not os.path.exists(EXCEL_FILE):

            # Crear archivo Excel con la hoja indicada
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                df_new.to_excel(writer, sheet_name=sheet_name, index=False)

            return

        # Cargar libro existente
        book = load_workbook(EXCEL_FILE)

        # Si la hoja ya existe
        if sheet_name in book.sheetnames:

            # Leer datos actuales de la hoja
            df_old = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)

            # Añadir nuevo jugador a los datos existentes
            df_final = pd.concat([df_old, df_new], ignore_index=True)

        else:
            # Si la hoja no existe, usar solo el nuevo jugador
            df_final = df_new

        # Guardar hoja en el archivo Excel (reemplazando la anterior)
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_final.to_excel(writer, sheet_name=sheet_name, index=False)

    # Función auxiliar para eliminar jugador de Excel
    def delete_player(sheet_name, player_name):

        # Si el archivo no existe, no hacer nada
        if not os.path.exists(EXCEL_FILE):
            return

        # Cargar datos de la hoja
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)

        # Filtrar dataframe eliminando el jugador seleccionado
        df = df[df["Player"] != player_name]

        # Guardar nuevamente la hoja sin el jugador eliminado
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
    # Layout de la página
    form_col, col_viewer = st.columns([1, 2])
            
    # Formulario para añadir jugador
    with form_col:
        
        st.subheader("Add Player")

        # Selector de liga 
        league_options = ["All"] + sorted(list_df_players["stats_Comp"].unique())
        league = st.selectbox(
            "League",
            league_options,
            key="list_league_filter"
        )

        # Selector de equipo 
        equipos = sorted(
            list_df_players[list_df_players["stats_Comp"] == league]["stats_Squad"].unique()
        ) if league != "All" else sorted(list_df_players["stats_Squad"].unique())

        team = st.selectbox(
            "Team",
            ["All"] + equipos,
            key="list_team_filter"
        )

        # Filtrar jugadores por liga y equipo 
        jugadores_filtro = list_df_players.copy()

        if league != "All":
            jugadores_filtro = jugadores_filtro[jugadores_filtro["stats_Comp"] == league]

        if team != "All":
            jugadores_filtro = jugadores_filtro[jugadores_filtro["stats_Squad"] == team]

        # Selector de posición 
        posiciones_unicas = list(jugadores_filtro["stats_Pos"].unique())
        posiciones_disponibles = [p for p in pos_order if p in posiciones_unicas]

        pos = st.selectbox(
            "Position",
            ["All"] + posiciones_disponibles,
            key="list_pos_filter"
        )

        if pos != "All":
            jugadores_filtro = jugadores_filtro[jugadores_filtro["stats_Pos"] == pos]

        # Selector de jugador 
        player_to_add = st.selectbox(
            "Select Player",
            ["Select"] + list(jugadores_filtro["Player"]),
            key="list_player_to_add"
        )

        # Formulario de datos de scout
        with st.form("player_form"):
            
            st.write("Add Scout Data")
            
            # Selector de usuario que crea la lista
            user_selected = st.selectbox(
                "User",
                ["Select"] + list_users_list,
                key="list_user_select"
            )
            
            # Selector de tipo de lista de scouting
            list_choice = st.selectbox(
                "List",
                ["Select"] + LISTS,   # Opciones disponibles de listas
                key="list_list_choice"
            )

            # Campo de texto para añadir comentario opcional
            comment = st.text_area(
                "Comment (optional)",
                key="list_comment"
            )

            # Selector de valoración del scout
            note = st.radio(
                "Scout Rating",
                ["Target","Watch","Prospect"],   # Niveles de interés del jugador
                key="list_note"
            )

            # Botón para enviar el formulario
            submitted = st.form_submit_button("Add Player")

            # Lógica al enviar el formulario
            if submitted:

                # Validar que se haya seleccionado un jugador
                if league == "All":
                    st.warning("Please select a league.")
                elif team == "All":
                    st.warning("Please select a team.")
                elif pos == "All":
                    st.warning("Please select a position.")
                elif player_to_add == "Select":
                    st.warning("Please select a player.")    
                elif user_selected == "Select":
                    st.warning("Please select a user.")
                elif list_choice == "Select":
                    st.warning("Please select a list.")

                # Si pasa todas las validaciones, guardar
                else:

                    # Crear diccionario con los datos del jugador
                    player_data = {

                        "Player": player_to_add,
                        "User": user_selected,
                        "League": league,
                        "Team": team,
                        "Position": pos,
                        "List": list_choice,
                        "Comment": comment,
                        "Note": note
                    }

                    # Guardar jugador en el archivo Excel
                    save_player_to_excel(player_data, list_choice)

                    # Confirmación visual al usuario
                    st.success(f"Player saved: {player_to_add}")
    
    # Botón Reset 
    if st.button("🔄 Reset Filters"):

        st.session_state["list_do_reset"] = True
        st.rerun()
        
    # Visualización de listas
    with col_viewer:

        # Título de la sección donde se visualizan las listas
        st.subheader("Player Lists")

        # Crear columnas para los filtros de visualización
        list_col1, list_col2 = st.columns(2)

        with list_col1:

            # Selector de lista de scouting
            selected_list = st.selectbox(
                "Select list",
                ["All"] + LISTS,   # Permite ver todas o una lista específica
                key="list_view_list"
            )

        with list_col2:

            # Selector de usuario que creó la lista
            usar_list = st.selectbox(
                "User list",
                ["All"] + list_users_list,   # Filtrar por usuario
                key="list_view_users_list"
            )

        # Si se selecciona "All", cargar todas las listas
        if selected_list == "All":

            dfs = []

            # Cargar cada lista por separado
            for l in LISTS:

                df_temp = load_sheet(l)
                df_temp["List"] = l   # Añadir columna con nombre de la lista
                dfs.append(df_temp)

            # Unir todas las listas en un único dataframe
            df_list = pd.concat(dfs, ignore_index=True)

        else:

            # Cargar solo la lista seleccionada
            df_list = load_sheet(selected_list)

        # Filtrar por usuario si se selecciona uno específico
        if usar_list != "All":

            df_list = df_list[df_list["User"] == usar_list]

        # Mostrar mensaje si no hay jugadores
        if df_list.empty:

            st.info("No players in this list")

        else:
            
            # Limitar la muestra a 8 jugadores por página
            display_limit = 8
            df_display = df_list.head(display_limit)
            
            # Cantidad de jugadores por página
            display_limit = 8

            # Inicializar página actual
            if "page_idx" not in st.session_state:
                st.session_state["page_idx"] = 0

            # Calcular número total de páginas
            total_pages = (len(df_list) - 1) // display_limit + 1

            # Botones de navegación
            nav_col1, nav_col2, nav_col3 = st.columns([1, 4, 1])

            with nav_col1:
                if st.button("⬅️ Previous") and st.session_state["page_idx"] > 0:
                    st.session_state["page_idx"] -= 1

            with nav_col3:
                if st.button("Next ➡️") and st.session_state["page_idx"] < total_pages - 1:
                    st.session_state["page_idx"] += 1

            # Calcular rango a mostrar según página
            start_idx = st.session_state["page_idx"] * display_limit
            end_idx = start_idx + display_limit
            df_display = df_list.iloc[start_idx:end_idx]

            # Mostrar mensaje si no hay jugadores
            if df_display.empty:
                st.info("No players in this page")
            else:
                # Contenedor para jugadores
                for i, row in df_display.iterrows():
                    player_register_c1, player_register_c2 = st.columns([12, 1])
                    with player_register_c1:
                        st.markdown(
                            f"""
                            **{row['Player']}** | 🏟️ {row['Team']} | 🏆 {row['League']}  
                            📝 {row['Comment']}  
                            ⭐ **{row['Note']}** | 👤 **{row['User']}**
                            """
                        )
                    with player_register_c2:
                        if st.button("🗑️", key=f"del_{row['Player']}_{start_idx+i}"):
                            delete_player(row["List"], row["Player"])
                            st.rerun()

            # Info de la página
            st.info(f"Page {st.session_state['page_idx'] + 1} of {total_pages}")

    # Botón para imprimir la página
    components.html(
    """
    <button onclick="parent.window.print()" style="
        padding:8px 14px;
        font-size:16px;
        cursor:pointer;
        background-color:#2563eb;
        color:white;
        border:none;
        border-radius:6px;">
        🖨️ Print Page
    </button>
    """,
    height=55
    )
    
    if not df_list.empty:

        if st.button("⚙️ Prepare PDF"):

            pdf = FPDF()
            pdf.add_page()

            pdf.add_font("DejaVu", "", str(ASSETSFONTS / "DejaVuSans.ttf"), uni=True)
            pdf.set_font("DejaVu", "", 25)

            pdf.cell(0, 10, "List of Players", ln=True, align="C")
            pdf.ln(5)

            col_widths = [35, 25, 25, 20, 20, 30, 20]
            headers = ["Player", "Team", "League", "Position", "List", "Note", "User"]

            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align="C")
            pdf.ln()

            def truncate(text, max_len=10):
                return text[:max_len] + "..." if len(text) > max_len else text

            pdf.set_font("DejaVu", "", 7)
            for _, row in df_list.iterrows():
                pdf.cell(col_widths[0], 8, truncate(str(row["Player"])), border=1)
                pdf.cell(col_widths[1], 8, truncate(str(row["Team"])), border=1)
                pdf.cell(col_widths[2], 8, truncate(str(row["League"])), border=1)
                pdf.cell(col_widths[3], 8, str(row["Position"]), border=1)
                pdf.cell(col_widths[4], 8, str(row["List"]), border=1)
                pdf.cell(col_widths[5], 8, truncate(str(row["Note"])), border=1)
                pdf.cell(col_widths[6], 8, str(row["User"]), border=1)
                pdf.ln()

            # Logo
            logo_buffer = get_watermark(alpha=10)
            pdf.image(logo_buffer, x=55, y=100, w=100)

            # Generar PDF en memoria correctamente
            pdf_bytes = pdf.output(dest="S").encode("latin-1")

            b64 = base64.b64encode(pdf_bytes).decode()

            components.html(
                f"""
                <a href="data:application/pdf;base64,{b64}" download="player_list.pdf">
                    <button style="
                        padding:8px 14px;
                        font-size:14px;
                        cursor:pointer;
                        background-color:#dc2626;
                        color:white;
                        border:none;
                        border-radius:6px;">
                        📄 Export to PDF
                    </button>
                </a>
                """,
                height=55
            )